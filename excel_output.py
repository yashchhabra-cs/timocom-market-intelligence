import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

OUTPUT_DIR = "output"
OUTPUT_FILE = "market_intelligence.xlsx"

HEADERS = ["Company", "URL", "Summary", "Key Insights", "Trends", "Recommendations"]

HEADER_FONT = Font(name="Calibri", bold=True, color="000000")
CONTENT_FONT = Font(name="Calibri", color="00008B")
HEADER_FILL = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
WRAP = Alignment(wrap_text=True, vertical="top")

COLUMN_WIDTHS = {
    "Company": 20,
    "URL": 35,
    "Summary": 45,
    "Key Insights": 45,
    "Trends": 45,
    "Recommendations": 45,
}


def cell_value(value):
    if isinstance(value, list):
        return "\n".join(f"- {item}" for item in value)
    if value is None:
        return ""
    return str(value)


def save_to_excel(results):
    wb = Workbook()
    ws = wb.active
    ws.title = "Market Intelligence"

    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = WRAP
        ws.column_dimensions[get_column_letter(col_idx)].width = COLUMN_WIDTHS[header]

    ws.row_dimensions[1].height = 20

    for row_idx, company_result in enumerate(results, start=2):
        company = company_result.get("company", "")
        url = company_result.get("url", "")
        analysis = company_result.get("analysis", {})

        if isinstance(analysis, dict) and "error" in analysis:
            summary = f"ERROR: {analysis['error']}"
            key_insights = ""
            trends = ""
            recommendations = ""
        else:
            try:
                summary = cell_value(analysis.get("summary"))
            except Exception:
                summary = "(error reading summary)"

            try:
                key_insights = cell_value(analysis.get("key_insights"))
            except Exception:
                key_insights = "(error reading key insights)"

            try:
                trends = cell_value(analysis.get("trends"))
            except Exception:
                trends = "(error reading trends)"

            try:
                recommendations = cell_value(analysis.get("recommendations"))
            except Exception:
                recommendations = "(error reading recommendations)"

        row_values = [company, url, summary, key_insights, trends, recommendations]

        for col_idx, value in enumerate(row_values, start=1):
            try:
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = CONTENT_FONT
                cell.alignment = WRAP
            except Exception as e:
                print(f"  [WARNING] Could not write cell ({row_idx}, {col_idx}): {e}")

        ws.row_dimensions[row_idx].height = 80

    try:
        os.makedirs(OUTPUT_DIR, exist_ok=True)
    except OSError as e:
        print(f"[ERROR] Could not create output directory '{OUTPUT_DIR}': {e}")
        return

    output_path = os.path.join(OUTPUT_DIR, OUTPUT_FILE)

    try:
        wb.save(output_path)
        print(f"[SUCCESS] Excel file saved to: {os.path.abspath(output_path)}")
    except PermissionError:
        print(f"[ERROR] Permission denied saving to '{output_path}'. Close the file if it is open.")
    except Exception as e:
        print(f"[ERROR] Failed to save Excel file: {e}")


if __name__ == "__main__":
    from scraper import scrape_websites, WEBSITES
    from processor import process_results

    scraped_data = scrape_websites(WEBSITES)
    results = process_results(scraped_data)
    save_to_excel(results)
